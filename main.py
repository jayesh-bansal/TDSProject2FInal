from fastapi import FastAPI, Form, File, UploadFile  # type: ignore
from fastapi.responses import HTMLResponse  # type: ignore
from fastapi.middleware.cors import CORSMiddleware  # type: ignore
import os
import openpyxl  # type: ignore
from processing import fetch_answer
import re
import stat

app = FastAPI()

# CORS Configuration (Vercel allows any origin by default)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "tasks.xlsx")


def load_tasks_from_excel():
    if not os.path.exists(EXCEL_FILE):
        return {}
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    tasks = {row[0]: row[1] for row in sheet.iter_rows(
        min_row=2, values_only=True) if row[0] and row[1]}

    tasks_answers = {row[0]: row[2] for row in sheet.iter_rows(
        min_row=2, values_only=True) if row[0] and row[2]}
    workbook.close()
    return (tasks, tasks_answers)


TASKS, TASKS_ANSWERS = load_tasks_from_excel()


def classify_task(question: str) -> str:
    """Classify a question based on keyword matching with TASKS."""
    question_lower = question.lower()  # Convert to lowercase for case-insensitive matching
    for task_id, keywords in TASKS.items():
        if any(keyword.lower() in question_lower for keyword in keywords.split(",")):
            return task_id  # Return the first matching task ID
    return "Unknown"  # Default if no match is found


def save_file(file: UploadFile):
    os.makedirs("uploads", exist_ok=True)
    if not file or not file.filename:
        return "Error: No file provided."
    # Define the file path
    file_path = os.path.join(os.getcwd(), "uploads", file.filename)
    try:
        # Write the file content manually
        with open(file_path, "wb") as buffer:
            buffer.write(file.file.read())
        # Set file permissions to 777 (read, write, execute for all)
        os.chmod(file_path, stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
    except Exception as e:
        return f"Error saving file: {str(e)}"
    return file_path


def get_file_path(question: str) -> str:
    """Extracts a single filename from the question and returns its full path in the /uploads directory."""
    match = re.search(r'([^\/\\\s]+?\.[a-zA-Z0-9]+)', question)
    file = match.group(1) if match else None
    file_path = os.path.join(os.getcwd(), "uploads", file) if file else None
    return file_path if file_path and os.path.exists(file_path) else None


@app.get("/", response_class=HTMLResponse)
async def serve_form():
    file_path = os.path.join(os.path.dirname(__file__), "index.html")
    try:
        with open(file_path, "r") as file:
            return HTMLResponse(content=file.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>index.html not found</h1>", status_code=404)

async def read_answer(task_id: str, question: str):
    print("reading from json")
    answer = TASKS_ANSWERS.get(task_id, "No answer found for this task.")
    return answer

@app.post("/api/")
async def receive_question(question: str = Form(...), file: UploadFile = File(None)):
    # if 'where is ' in question.lower():
    #     file_path = get_file_path(question)
    #     return {"question": question, "answer": file_path if file_path else "File not found"}

    task_id = classify_task(question)

    if task_id in ['GA1.2', 'GA1.4', 'GA1.5', 'GA1.7', 'GA1.9', 'GA1.18']:
            answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ['GA1.6', 'GA1.11']:
        func_answer=""
        if file:
            func_answer = await fetch_answer(task_id=task_id, question=question, file_path=file_path)
        answer = func_answer or await read_answer(task_id=task_id,question=question)
    elif task_id in ['GA1.3', 'GA1.8', 'GA1.10', 'GA1.12','GA1.14','GA1.15', 'GA1.16', 'GA1.17']:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        else:
            answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA2.2','GA2.4']:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        else:
            answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA2.5']:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        else:
            answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ['GA2.9']:
        if file:
            file_path = save_file(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file_path)
        else:
            answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ["GA3.1", "GA3.2", "GA3.3","GA3.5","GA3.6"]:
        answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ["GA3.4"]:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
    elif task_id in ['GA5.1','GA5.2','GA5.3','GA5.4','GA5.5','GA5.6','GA5.7']:
        if file:
            file_path = save_file(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file_path)
        else:
            answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA5.8']:
        answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ['GA5.10']:
        if file:
            file_path = save_file(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file_path)
            return {"answer": answer}
    else:
        if file:
            file_path = save_file(file)
        answer = await read_answer(task_id=task_id, question=question)

    if isinstance(answer, int):
        answer = str(answer)
    if isinstance(answer, float):
        answer = str(answer)
    if isinstance(answer, list):
        answer = str(answer)
    
    output = {"question": question,"task": task_id,"answer": answer,"file received": file.filename if file else "No file uploaded", }
    print(output)
    print()
    return {"answer": answer}
